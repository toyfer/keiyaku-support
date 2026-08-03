"""市様式 Excel（templates_excel）への簡易マッピング

公式大ひな形とは別に、入札書・委任状・見積書など単票を埋める。
1) {{placeholder}} があれば置換
2) FORM_CELL_MAPS に座標があればセル書き込み
座標は実ファイルで調整可能（docs/FORM_CELL_MAPS.md）。
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

from .config import OUTPUT_DIR, TEMPLATES_EXCEL_DIR

_PLACEHOLDER = re.compile(r"\{\{\s*(\w+)\s*\}\}")

# ファイル名の部分一致 → 論理キー
FORM_FILE_KEYS = {
    "見積書【市様式　業務用】": "見積書_業務",
    "見積書【市様式　物品用】": "見積書_物品",
    "入札書【市様式　業務用】": "入札書_業務",
    "入札書【市様式　物品用】": "入札書_物品",
    "委任状【市様式　業務用】": "委任状_業務",
    "委任状【市様式　物品用】": "委任状_物品",
    "入札辞退届【業務用】": "辞退届_業務",
    "入札辞退届【物品用】": "辞退届_物品",
    "見積り合せ辞退届【業務用】": "見積辞退_業務",
    "見積り合せ辞退届【物品用】": "見積辞退_物品",
}

# 論理キー → {cell: value_key}  （運用で座標を埋める。空でもプレースホルダは動く）
FORM_CELL_MAPS: dict[str, dict[str, str]] = {
    "見積書_業務": {
        # 例: "B5": "title", "B10": "supplier_name",
    },
    "見積書_物品": {},
    "入札書_業務": {},
    "入札書_物品": {},
    "委任状_業務": {},
    "委任状_物品": {},
}


def list_city_form_files() -> list[Path]:
    if not TEMPLATES_EXCEL_DIR.exists():
        return []
    return sorted(TEMPLATES_EXCEL_DIR.glob("*.xlsx"))


def resolve_city_form(key_or_name: str) -> Path | None:
    """論理キーまたはファイル名部分一致で解決。"""
    files = list_city_form_files()
    if not files:
        return None
    # exact key
    for path in files:
        for frag, key in FORM_FILE_KEYS.items():
            if frag in path.name and (key_or_name == key or key_or_name in path.name or key_or_name == frag):
                return path
    for path in files:
        if key_or_name in path.name:
            return path
    return None


def logical_key_for(path: Path) -> str | None:
    for frag, key in FORM_FILE_KEYS.items():
        if frag in path.name:
            return key
    return None


def fill_city_form(src: str | Path, values: dict, out_dir: str | Path | None = None,
                   output_name: str | None = None) -> Path:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    src = Path(src)
    out_dir = Path(out_dir or OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    dst = out_dir / (output_name or src.name)
    shutil.copy2(src, dst)

    wb = load_workbook(dst)
    key = logical_key_for(src) or ""
    cell_map = FORM_CELL_MAPS.get(key) or {}

    def set_cell(ws, addr, value):
        cell = ws[addr]
        if isinstance(cell, MergedCell):
            for mr in ws.merged_cells.ranges:
                if addr in mr:
                    addr = f"{get_column_letter(mr.min_col)}{mr.min_row}"
                    cell = ws[addr]
                    break
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return
        cell.value = value

    # 座標マップ
    for sheet in wb.worksheets:
        for addr, vkey in cell_map.items():
            if vkey in values and values[vkey] not in (None, ""):
                try:
                    set_cell(sheet, addr, values[vkey])
                except Exception:
                    pass

    # プレースホルダ
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    cell.value = _PLACEHOLDER.sub(
                        lambda m: str(values.get(m.group(1), "") or ""), cell.value
                    )

    wb.save(dst)
    wb.close()
    return dst


def generate_recommended_city_forms(case: dict, values: dict, method: str | None = None,
                                    case_type: str | None = None) -> list[Path]:
    from .template_profile import recommended_city_forms

    method = method or case.get("contract_method")
    case_type = case_type or case.get("case_type")
    outs = []
    safe_no = str(case.get("case_no") or "case").replace("/", "-")
    for key in recommended_city_forms(method, case_type):
        path = resolve_city_form(key)
        if not path:
            continue
        out_name = f"{safe_no}_{path.name}"
        try:
            outs.append(fill_city_form(path, values, output_name=out_name))
        except Exception:
            continue
    return outs
