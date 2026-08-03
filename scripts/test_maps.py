"""変更マップがひな形ラベルと一致するか（パッケージB必須）"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config import DATA_DIR, TEMPLATES_DIR
from app.fields import CHANGE_MAX, change_map, formula_cells_for


def _find(name: str) -> Path | None:
    for base in (TEMPLATES_DIR, DATA_DIR / "templates", ROOT / "data" / "templates"):
        p = base / name
        if p.exists():
            return p
    return None


def labels(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)
    if "入力シート" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"no 入力シート in {path}")
    ws = wb["入力シート"]
    out = {}
    for r in range(1, 160):
        lab = ws.cell(r, 9).value
        q = ws.cell(r, 17).value
        is_f = isinstance(q, str) and str(q).startswith("=")
        if lab or q is not None:
            out[r] = (str(lab or "").strip(), "F" if is_f else "V")
    wb.close()
    return out


def check(case_type: str, filename: str) -> bool:
    path = _find(filename)
    if not path:
        print(f"SKIP {filename} (place package B under data/templates/)")
        return True
    labs = labels(path)
    print(f"== {filename} ({case_type}) ==")
    ok = True
    for n in range(1, CHANGE_MAX + 1):
        cm = change_map(case_type, n)
        start = int(cm["change_notice"][1:])
        lab, _ = labs.get(start, ("?", "?"))
        if "協議通知" not in lab:
            print(f"  FAIL change#{n} Q{start} label={lab!r}")
            ok = False
        else:
            print(f"  OK change#{n} Q{start} ({lab})")
    print(f"  formula cells: {len(formula_cells_for(case_type))}")
    return ok


def main():
    assert change_map("工事", 1)["change_notice"] == "Q61"
    assert change_map("工事", 2)["change_notice"] == "Q78"
    assert change_map("業務", 1)["change_notice"] == "Q67"
    assert change_map("業務", 5)["change_notice"] == "Q135"
    ok1 = check("工事", "hinagata_koji.xlsx")
    ok2 = check("業務", "hinagata_gyomu.xlsx")
    print("PASS" if ok1 and ok2 else "FAIL")
    sys.exit(0 if ok1 and ok2 else 1)


if __name__ == "__main__":
    main()
