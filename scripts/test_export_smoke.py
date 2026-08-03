"""デモDB + ひな形B があれば変更3回生成スモーク"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config import DB_PATH, TEMPLATES_DIR
from app.db import init_db
from app.case_service import build_export_values, get_case, list_cases, list_changes, save_change
from app.document_engine import generate_from_case, resolve_template
from app.fields import change_map


def main():
    if not resolve_template("工事") and not resolve_template("業務"):
        print("SKIP: no hinagata in data/templates/ (package B)")
        return
    init_db()
    if not DB_PATH.exists():
        print("SKIP: no DB")
        return
    cases = list_cases()
    if not cases:
        print("SKIP: no cases in DB (need package D seed)")
        return
    koji = next((c for c in cases if c.get("case_type") == "工事"), cases[0])
    cid = koji["id"]
    print("case", koji.get("case_no"), koji.get("title"))
    for n, delta in [(1, 1_100_000), (2, 550_000), (3, -220_000)]:
        save_change(cid, {
            "change_no": n,
            "change_notice": f"令和7年{6+n}月1日",
            "change_explain": f"令和7年{6+n}月2日 10:00",
            "change_kessai_date": f"令和7年{6+n}月3日",
            "change_date": f"令和7年{6+n}月4日",
            "change_dir": "増" if delta > 0 else "減",
            "amount_delta": delta,
            "new_end_date": "令和7年12月28日",
            "budget_allocated": 80_000_000,
            "budget_executed": 10_000_000 + n * 1000,
            "status": "applied",
        })
    changes = list_changes(cid)
    case = get_case(cid)
    values = build_export_values(case, changes)
    out_dir = Path(tempfile.mkdtemp(prefix="keiyaku_out_"))
    out = generate_from_case(case, values, out_dir=out_dir, changes=changes)
    print("out", out, out.stat().st_size)

    from openpyxl import load_workbook
    wb = load_workbook(out, data_only=False)
    ws = wb["入力シート"]
    for n in (1, 2, 3):
        cm = change_map("工事" if case.get("case_type") == "工事" else "業務", n)
        row = int(cm["change_dir"][1:])
        val = ws.cell(row, 17).value
        print(f"  change#{n} Q{row}={val!r}")
        assert val in ("増", "減"), val
    cm4 = change_map("工事", 4)
    row4 = int(cm4["change_dir"][1:])
    print(f"  change#4 cleared? Q{row4}={ws.cell(row4, 17).value!r}")
    q18 = ws.cell(18, 17).value
    if isinstance(q18, str) and q18.startswith("="):
        print("  Q18 formula OK", q18[:40])
    wb.close()
    print("export smoke OK")


if __name__ == "__main__":
    main()
